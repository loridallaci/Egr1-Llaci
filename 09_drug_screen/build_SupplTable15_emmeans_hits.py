"""
Build Supplementary Table 15 - confirmation-screen emmeans results.

Source : 09_drug_screen/confirmation_screen_sexdiff_emmeans/output/emmeans_sexdiff_results.xlsx
Output : 09_drug_screen/SupplementaryTable15_emmeans_sexbiased_hits.xlsx

Combines the three emmeans contrasts into one table:
  Q2 WT  = (male - female | Egr1 WT)
  Q2 KD  = (male - female | Egr1 KD)
  Q4     = (Q2 WT) - (Q2 KD), the Egr1 x sex interaction
A compound is listed if ANY contrast has raw P <= 0.05. Negative estimate = male
more sensitive. BH-adjusted q reported alongside but hits are called on raw P.

The source workbook cannot be opened with openpyxl (missing drawing part), so it
is parsed as a zip of sheet XML.
"""
import zipfile, os
from xml.etree import ElementTree as ET
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

NS = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
HERE = os.path.dirname(os.path.abspath(__file__))
SRC  = os.path.join(HERE, "confirmation_screen_sexdiff_emmeans", "output",
                    "emmeans_sexdiff_results.xlsx")
OUT  = os.path.join(HERE, "SupplementaryTable15_emmeans_sexbiased_hits.xlsx")
ALPHA = 0.05


def read_sheet(path, want):
    z = zipfile.ZipFile(path)
    shared = [''.join(t.text or '' for t in si.iter(NS + 't'))
              for si in ET.fromstring(z.read('xl/sharedStrings.xml')).iter(NS + 'si')]
    wbx = ET.fromstring(z.read('xl/workbook.xml'))
    rels = {r.get('Id'): r.get('Target')
            for r in ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))}
    for sh in wbx.iter(NS + 'sheet'):
        if sh.get('name') != want:
            continue
        t = rels[sh.get('{http://schemas.openxmlformats.org/officeDocument/2006/'
                        'relationships}id')].lstrip('/')
        if not t.startswith('xl/'):
            t = 'xl/' + t
        rows = []
        for row in ET.fromstring(z.read(t)).iter(NS + 'row'):
            v = []
            for c in row.iter(NS + 'c'):
                e = c.find(NS + 'v')
                v.append((shared[int(e.text)] if c.get('t') == 's' else e.text)
                          if e is not None else None)
            if any(x is not None for x in v):
                rows.append(v)
        return rows
    raise KeyError(want)


rows = read_sheet(SRC, 'emmeans_results')
hdr = [str(x) for x in rows[0]]
I = {k: hdr.index(k) for k in hdr}
num = lambda r, k: (float(r[I[k]]) if r[I[k]] not in (None, '') else None)

# disambiguate duplicate compound names (THIOGUANINE appears twice)
names = [r[I['Compound']] for r in rows[1:]]
seen = {}
label = []
for n in names:
    if names.count(n) > 1:
        seen[n] = seen.get(n, 0) + 1
        label.append("%s (run %d)" % (n.title(), seen[n]))
    else:
        label.append(n.title())

recs = []
for lab, r in zip(label, rows[1:]):
    wt_p, kd_p, q4_p = num(r, 'Q2WT_p'), num(r, 'Q2KD_p'), num(r, 'Q4_p')
    wt_e, kd_e = num(r, 'Q2WT_est'), num(r, 'Q2KD_est')
    wt_hit, kd_hit, dep = wt_p <= ALPHA, kd_p <= ALPHA, q4_p <= ALPHA
    if wt_hit and kd_hit:      cat = "Sex-biased in both"
    elif wt_hit:               cat = "WT only"
    elif kd_hit:               cat = "KD only"
    else:                      cat = "Not sex-biased"
    if dep and wt_hit and not kd_hit:      fate = "Abolished by Egr1 KD"
    elif dep:                              fate = "Reduced by Egr1 KD"
    elif wt_hit or kd_hit:                 fate = "Egr1-independent"
    else:                                  fate = ""
    pct = (100 * (1 - abs(kd_e) / abs(wt_e))) if (wt_e and abs(wt_e) > 0) else None
    recs.append(dict(
        Compound=lab, Class=r[I['Class']],
        WT_estimate=wt_e, WT_SE=num(r, 'Q2WT_SE'), WT_t=num(r, 'Q2WT_t'),
        WT_P=wt_p, WT_q=num(r, 'Q2WT_q'),
        WT_hit="Yes" if wt_hit else "No",
        WT_direction=(r[I['Q2WT_sex']] + "-biased") if wt_hit else "",
        KD_estimate=kd_e, KD_SE=num(r, 'Q2KD_SE'), KD_t=num(r, 'Q2KD_t'),
        KD_P=kd_p, KD_q=num(r, 'Q2KD_q'),
        KD_hit="Yes" if kd_hit else "No",
        KD_direction=(r[I['Q2KD_sex']] + "-biased") if kd_hit else "",
        Interaction_estimate=num(r, 'Q4_est'), Interaction_SE=num(r, 'Q4_SE'),
        Interaction_t=num(r, 'Q4_t'), Interaction_P=q4_p, Interaction_q=num(r, 'Q4_q'),
        Egr1_dependent="Yes" if dep else "No",
        Category=cat, Outcome=fate,
        Pct_reduction_in_sex_difference=(round(pct, 1) if pct is not None else None),
        df=num(r, 'df')))

hits = [x for x in recs if x['Category'] != "Not sex-biased"]
hits.sort(key=lambda x: (x['Egr1_dependent'] != "Yes", x['Interaction_P']))
recs_sorted = sorted(recs, key=lambda x: x['WT_P'])

COLS = list(recs[0].keys())
HEAD = Font(bold=True, color="FFFFFF")
FILL = PatternFill("solid", fgColor="3B4CC0")

wb = openpyxl.Workbook()

ws = wb.active
ws.title = "README"
readme = [
 ["Supplementary Table 15. Sex-biased drug responses in the 85-compound confirmation screen."],
 [""],
 ["Source", "09_drug_screen/confirmation_screen_sexdiff_emmeans/output/emmeans_sexdiff_results.xlsx"],
 ["Built by", "09_drug_screen/build_SupplTable15_emmeans_hits.py"],
 [""],
 ["Model", "Per compound, 12 well-level Z-scores (3 replicates x 4 genotypes) were fitted as"],
 ["", "z ~ sex * genotype by linear regression (residual d.f. = 8). Estimated marginal means"],
 ["", "were used to form three contrasts, tested by two-sided t with the pooled residual error."],
 [""],
 ["Q2 WT", "(male - female) within Egr1 WT cells. Negative = male more sensitive."],
 ["Q2 KD", "(male - female) within Egr1 KD cells."],
 ["Interaction", "(Q2 WT) - (Q2 KD). Negative = sex difference larger in WT, i.e. reduced by Egr1 KD."],
 [""],
 ["Hit definition", "raw P <= 0.05 on the relevant contrast. BH-adjusted q is reported alongside"],
 ["", "but was not used to call hits."],
 ["Egr1-dependent", "interaction contrast raw P <= 0.05."],
 [""],
 ["Sheet 'Sex-biased hits'", "compounds significant in at least one Q2 contrast (n = 64 rows)."],
 ["Sheet 'All compounds'", "all 85 compounds, including the 2 negative controls and non-hits."],
 [""],
 ["Note - THIOGUANINE", "present twice in the compound plate; both runs retained and labelled"],
 ["", "(run 1) / (run 2). This is why the union of WT and KD hits is 64 rows / 63 compounds."],
 ["Note - controls", "SECNIDAZOLE and METHYLPHENIDATE HYDROCHLORIDE were negative controls;"],
 ["", "neither is a hit."],
 ["Note - denominators", "63 compounds sex-biased in WT; 53 in KD; 64 in either; 37 Egr1-dependent."],
]
for r in readme:
    ws.append(r)
ws["A1"].font = Font(bold=True, size=12)
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 105
for row in ws.iter_rows():
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)

def add(title, data):
    s = wb.create_sheet(title)
    s.append(COLS)
    for c in s[1]:
        c.font = HEAD; c.fill = FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for d in data:
        s.append([d[c] for c in COLS])
    for i, c in enumerate(COLS, 1):
        s.column_dimensions[get_column_letter(i)].width = \
            max(11, min(30, len(c) + 3)) if c not in ("Compound", "Class") else 30
    s.freeze_panes = "C2"
    for row in s.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = '0.000E+00' if abs(cell.value) < 1e-3 and cell.value != 0 else '0.000'
    return s

add("Sex-biased hits", hits)
add("All compounds", recs_sorted)
wb.save(OUT)

nWT = sum(1 for x in recs if x['WT_hit'] == "Yes")
nKD = sum(1 for x in recs if x['KD_hit'] == "Yes")
nDep = sum(1 for x in recs if x['Egr1_dependent'] == "Yes")
print("compounds            : %d" % len(recs))
print("sex-biased in WT     : %d" % nWT)
print("sex-biased in KD     : %d" % nKD)
print("union (any contrast) : %d" % len(hits))
print("Egr1-dependent       : %d" % nDep)
print("  abolished          : %d" % sum(1 for x in hits if x['Outcome'] == "Abolished by Egr1 KD"))
print("  reduced            : %d" % sum(1 for x in hits if x['Outcome'] == "Reduced by Egr1 KD"))
print("  Egr1-independent   : %d" % sum(1 for x in hits if x['Outcome'] == "Egr1-independent"))
print("\nwrote %s" % OUT)
